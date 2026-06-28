#!/usr/bin/env python3
"""
Парсер Яндекс.Карты + 2ГИС
=========================
Сбор данных о компаниях в Excel.
Поддерживает несколько городов и сфер, ночной режим, историю.

Запуск: python3 ymaps_parser.py
"""

import re, os, sys, time, random
from datetime import date, datetime
from difflib import get_close_matches
from urllib.parse import quote

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print(" ❌  pip install playwright")
    print("     python3 -m playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print(" ❌  pip install openpyxl")
    sys.exit(1)


# ══════════════════════════════════════════
#  ПУТИ BRAVE
# ══════════════════════════════════════════

BRAVE_PATH = "/opt/brave-bin/brave"
BRAVE_PROFILE = "/home/vi/.config/BraveSoftware/Brave-Browser/Default"


# ══════════════════════════════════════════
#  СПРАВОЧНИК
# ══════════════════════════════════════════

KNOWN_CITIES = {
    "ижевск":"Ижевск","воткинск":"Воткинск","глазов":"Глазов",
    "сарапул":"Сарапул","можга":"Можга","ува":"Ува",
    "камбарка":"Камбарка","игра":"Игра","балезино":"Балезино",
    "кез":"Кез","завьялово":"Завьялово","малая пурга":"Малая Пурга",
    "алнаши":"Алнаши","вавож":"Вавож","грахово":"Грахово",
    "дебесы":"Дебесы","киясово":"Киясово","красногорское":"Красногорское",
    "селы":"Селты","сюмси":"Сюмси","шаркан":"Шаркан",
    "юкаменское":"Юкаменское","яр":"Яр",
    "казань":"Казань","пермь":"Пермь","екатеринбург":"Екатеринбург",
    "киров":"Киров","москва":"Москва","санкт-петербург":"Санкт-Петербург",
}

CITY_SLUGS = {
    "ижевск":"izhevsk","воткинск":"votkinsk","глазов":"glazov",
    "сарапул":"sarapul","можга":"mozhga","ува":"uva",
    "камбарка":"kambarka","игра":"igra","балезино":"balezino",
    "кез":"kez","завьялово":"zavyalovo","малая пурга":"malaya-purga",
    "алнаши":"alnashi","вавож":"vavozh","грахово":"grakhovo",
    "дебесы":"debesy","киясово":"kiyasovo","красногорское":"krasnogorskoe",
    "селы":"selty","сюмси":"syumsi","шаркан":"sharkan",
    "юкаменское":"yukamenskoe","яр":"yar",
    "казань":"kazan","пермь":"perm","екатеринбург":"ekaterinburg",
    "киров":"kirov","москва":"moscow","санкт-петербург":"spb",
}


# ══════════════════════════════════════════
#  ЦВЕТА
# ══════════════════════════════════════════

class C:
    R="\033[91m"; G="\033[92m"; Y="\033[93m"; B="\033[94m"
    M="\033[95m"; Cc="\033[96m"; D="\033[90m"; N="\033[0m"
    BD="\033[1m"


# ══════════════════════════════════════════
#  ВВОД
# ══════════════════════════════════════════

def ask(prompt, default=""):
    if default:
        val = input(f"{prompt} [{C.D}{default}{C.N}]: ").strip()
    else:
        val = input(f"{prompt}: ").strip()
    return val if val else default


def resolve_city(raw: str) -> str:
    key = raw.lower().strip()
    if key in KNOWN_CITIES:
        return KNOWN_CITIES[key]
    matches = get_close_matches(key, KNOWN_CITIES.keys(), n=3, cutoff=0.7)
    if matches:
        variants = [KNOWN_CITIES[m] for m in matches]
        print(f"\n {C.Y}⚠  Город «{raw}» не найден. Возможно, вы имели в виду:{C.N}")
        for i, v in enumerate(variants, 1):
            print(f"   {i}. {v}")
        print(f"   {len(variants)+1}. Оставить «{raw}»")
        choice = ask("Выберите", str(len(variants)+1))
        try:
            n = int(choice)
            if 1 <= n <= len(variants):
                return variants[n-1]
        except ValueError:
            pass
    return raw


def get_2gis_slug(city_name: str) -> str:
    key = city_name.lower().strip()
    if key in CITY_SLUGS:
        return CITY_SLUGS[key]
    ru_to_en = {
        "а":"a","б":"b","в":"v","г":"g","д":"d","е":"e","ё":"e",
        "ж":"zh","з":"z","и":"i","й":"y","к":"k","л":"l","м":"m",
        "н":"n","о":"o","п":"p","р":"r","с":"s","т":"t","у":"u",
        "ф":"f","х":"kh","ц":"ts","ч":"ch","ш":"sh","щ":"shch",
        "ъ":"","ы":"y","ь":"","э":"e","ю":"yu","я":"ya",
    }
    slug = ""
    for ch in key:
        slug += ru_to_en.get(ch, ch if ch.isalnum() else "-")
    slug = re.sub(r"-+", "-", slug).strip("-")
    return slug if slug else city_name


# ══════════════════════════════════════════
#  ПАРСИНГ ЗАГОЛОВКА
# ══════════════════════════════════════════

def parse_title(title, platform):
    suffix = " — Яндекс" if "yandex" in platform.lower() else " — 2ГИС"
    clean = title.rsplit(suffix, 1)[0] if suffix in title else title
    parts = [p.strip() for p in clean.split(",")]
    name = parts[0] if parts else title
    category = parts[1] if len(parts) > 1 else ""
    address = ", ".join(parts[2:]) if len(parts) > 2 else ""
    return {"name": name, "category": category, "address": address}


# ══════════════════════════════════════════
#  ИЗВЛЕЧЕНИЕ ССЫЛОК
# ══════════════════════════════════════════

def extract_links(page, source):
    return page.evaluate(r"""(source) => {
        var siteLink = null, bookingLink = null, bookingExists = false;

        if (source === '2gis') {
            var links = document.querySelectorAll('a[href*="link.2gis"], a[href*="2gis.ru"]');
            links.forEach(function(link) {
                var text = (link.textContent || '').trim();
                var href = link.href || link.getAttribute('href') || '';

                if (/^[a-z0-9\u0430-\u044f][-a-z0-9\u0430-\u044f]*\.[a-z]{2,}/i.test(text)) {
                    if (!siteLink) {
                        siteLink = text.indexOf('http') === 0 ? text : 'https://' + text;
                    }
                }
                if (/запис|онлайн/i.test(text) && href.indexOf('link.2gis') > -1) {
                    if (!bookingLink) bookingLink = href;
                    bookingExists = true;
                }
            });
            return {site: siteLink, booking: bookingLink, booking_exists: bookingExists};
        }

        if (source === 'yandex') {
            var siteEl = document.querySelector('a.business-urls-view__link');
            if (siteEl && siteEl.href) {
                siteLink = siteEl.href;
            }
            document.querySelectorAll('a, button, [role="button"]').forEach(function(el) {
                var text = (el.textContent || '').trim().toLowerCase();
                if (/запис|онлайн/.test(text)) {
                    bookingExists = true;
                    var href = el.href || el.getAttribute('href') || '';
                    if (href && !bookingLink) bookingLink = href;
                }
            });
            return {site: siteLink, booking: bookingLink, booking_exists: bookingExists};
        }
        return {site: null, booking: null, booking_exists: false};
    }""", source)


# ══════════════════════════════════════════
#  ЗАДЕРЖКИ
# ══════════════════════════════════════════

def delay(night, a, b):
    """Спит от a до b секунд. В ночном режиме ×2."""
    m = 2 if night else 1
    time.sleep(random.uniform(a * m, b * m))


# ══════════════════════════════════════════
#  ПАРСИНГ ЯНДЕКС КАРТ
# ══════════════════════════════════════════

def parse_yandex(page, city, query, limit, seen, night):
    search_text = f"{query} {city}"
    url = f"https://yandex.ru/maps/?text={quote(search_text, safe='')}"

    page.goto(url, wait_until="domcontentloaded", timeout=35000)
    delay(night, 2.5, 4)

    for _ in range(3):
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        delay(night, 1.5, 2.5)

    org_links = page.query_selector_all('a[href*="/org/"]')
    urls = set()
    for link in org_links:
        href = link.get_attribute("href")
        if not href or "map/" in href:
            continue
        if href.startswith("/"):
            full = f"https://yandex.ru{href.split('?')[0]}"
        elif "yandex.ru" in href:
            full = href.split("?")[0]
        else:
            continue
        if not any(x in full for x in ["/gallery", "/reviews", "/prices", "#"]):
            urls.add(full)

    print(f"       {C.D}Найдено ссылок: {len(urls)}{C.N}")

    results = []
    for idx, org_url in enumerate(sorted(urls)):
        if len(results) >= limit:
            break

        delay(night, 2, 4)

        try:
            page.goto(org_url, wait_until="domcontentloaded", timeout=25000)
            delay(night, 1, 2)
        except Exception as e:
            print(f"       {C.Y}⚠  Ошибка: {e}{C.N}")
            continue

        html = page.content()
        title_m = re.search(r"<title>(.*?)</title>", html)
        raw_title = title_m.group(1).strip() if title_m else ""

        parsed = parse_title(raw_title, "yandex")
        name = parsed["name"]
        category = parsed["category"]
        address = parsed["address"]

        raw_phones = set(re.findall(r'\+7(\d{10})', html))
        phones = [f"+7 ({p[:3]}) {p[3:6]}-{p[6:8]}-{p[8:10]}" for p in sorted(raw_phones)]
        phones_str = ", ".join(phones)

        if not phones_str:
            continue

        key = (name.strip(), phones_str)
        if key in seen:
            continue
        seen.add(key)

        links = extract_links(page, "yandex")
        site = links.get("site") or "нет"
        booking = links.get("booking") or ""
        booking_exists = links.get("booking_exists", False)
        booking_val = booking if booking_exists else "Нет"
        if booking_exists and not booking:
            booking_val = org_url

        results.append({
            "source": "Яндекс.Карты",
            "category": category,
            "city": city,
            "name": name.strip(),
            "phones": phones_str,
            "booking": booking_val,
            "comment": address,
            "site": site,
            "platform_url": org_url,
        })
        print(f"       {C.G}✓{C.N} {name[:35]:35} {C.D}|{C.N} {phones_str[:20]} {C.D}|{C.N} {category[:20]}")

    return results


# ══════════════════════════════════════════
#  ПАРСИНГ 2ГИС
# ══════════════════════════════════════════

def parse_2gis(page, city, query, limit, seen, night):
    slug = get_2gis_slug(city)
    url = f"https://2gis.ru/{slug}/search/{quote(query)}"

    page.goto(url, wait_until="domcontentloaded", timeout=35000)
    delay(night, 3, 5)

    for _ in range(3):
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        delay(night, 1.5, 2.5)

    firm_links = page.query_selector_all('a[href*="/firm/"]')
    urls = set()
    for link in firm_links:
        href = link.get_attribute("href")
        if not href:
            continue
        clean = href.split("?")[0]
        full = f"https://2gis.ru{clean}" if clean.startswith("/") else clean
        if slug in full:
            urls.add(full)

    print(f"       {C.D}Найдено фирм: {len(urls)}{C.N}")

    results = []
    for idx, firm_url in enumerate(sorted(urls)):
        if len(results) >= limit:
            break

        delay(night, 2.5, 4.5)

        try:
            page.goto(firm_url, wait_until="domcontentloaded", timeout=25000)
            delay(night, 1.5, 2.5)
        except Exception as e:
            print(f"       {C.Y}⚠  Ошибка: {e}{C.N}")
            continue

        html = page.content()
        title_m = re.search(r"<title>(.*?)</title>", html)
        raw_title = title_m.group(1).strip() if title_m else ""

        parsed = parse_title(raw_title, "2gis")
        name = parsed["name"]
        category = parsed["category"]
        address = parsed["address"]

        raw_phones = set(re.findall(r'\+7(\d{10})', html))
        phones = [f"+7 ({p[:3]}) {p[3:6]}-{p[6:8]}-{p[8:10]}" for p in sorted(raw_phones)]
        phones_str = ", ".join(phones)

        if not phones_str:
            continue

        key = (name.strip(), phones_str)
        if key in seen:
            continue
        seen.add(key)

        links = extract_links(page, "2gis")
        site = links.get("site") or "нет"
        booking = links.get("booking") or ""
        booking_exists = links.get("booking_exists", False)
        booking_val = booking if booking_exists else "Нет"
        if booking_exists and not booking:
            booking_val = firm_url

        results.append({
            "source": "2ГИС",
            "category": category,
            "city": city,
            "name": name.strip(),
            "phones": phones_str,
            "booking": booking_val,
            "comment": address,
            "site": site,
            "platform_url": firm_url,
        })
        print(f"       {C.G}✓{C.N} {name[:35]:35} {C.D}|{C.N} {phones_str[:20]} {C.D}|{C.N} {category[:20]}")

    return results


# ══════════════════════════════════════════
#  ИСТОРИЯ
# ══════════════════════════════════════════

HISTORY_FILE = "history.xlsx"

def read_history():
    if not os.path.exists(HISTORY_FILE):
        return []
    try:
        wb = openpyxl.load_workbook(HISTORY_FILE)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                rows.append({
                    "city": str(row[0] or ""),
                    "sphere": str(row[1] or ""),
                    "count": int(row[2]) if row[2] else 0,
                    "date": str(row[3] or ""),
                    "file": str(row[4] or ""),
                })
        return rows
    except:
        return []


def append_history(city, sphere, count, filename):
    rows = read_history()
    rows.append({
        "city": city,
        "sphere": sphere,
        "count": count,
        "date": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "file": os.path.basename(filename),
    })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "История"
    headers = ["Город", "Сфера", "Записей", "Дата", "Файл"]
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill

    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r["city"])
        ws.cell(row=i, column=2, value=r["sphere"])
        ws.cell(row=i, column=3, value=r["count"])
        ws.cell(row=i, column=4, value=r["date"])
        ws.cell(row=i, column=5, value=r["file"])

    for col, w in enumerate([16, 22, 10, 18, 40], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(HISTORY_FILE)


def find_history(city, sphere, history):
    for h in history:
        if h["city"].lower() == city.lower() and h["sphere"].lower() == sphere.lower():
            return h
    return None


# ══════════════════════════════════════════
#  СОХРАНЕНИЕ
# ══════════════════════════════════════════

def save_excel(records, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Клиенты"

    headers = [
        "Сфера","Город","Наименование","Телефон",
        "Онлайн-запись","Комментарий","Ответственный",
        "Ссылка на сайт","Ссылка на Яндекс/2ГИС","В архиве","Дата создания"
    ]

    hf = Font(bold=True, size=11, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")

    today_s = date.today().strftime("%d.%m.%Y")
    for i, rec in enumerate(records, 2):
        ws.cell(row=i, column=1, value=rec.get("category", ""))
        ws.cell(row=i, column=2, value=rec["city"])
        ws.cell(row=i, column=3, value=rec["name"])
        ws.cell(row=i, column=4, value=rec["phones"])
        ws.cell(row=i, column=5, value=rec.get("booking", "Нет"))
        ws.cell(row=i, column=6, value=rec.get("comment", ""))
        ws.cell(row=i, column=7, value="")
        ws.cell(row=i, column=8, value=rec.get("site", "нет"))
        ws.cell(row=i, column=9, value=rec.get("platform_url", ""))
        ws.cell(row=i, column=10, value="")
        ws.cell(row=i, column=11, value=today_s)

    widths = [22, 14, 40, 38, 42, 60, 18, 50, 70, 10, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(filename)


# ══════════════════════════════════════════
#  СБОР ОДНОЙ ПАРЫ ГОРОД+СФЕРА
# ══════════════════════════════════════════

def collect_city_sphere(page, city, sphere, limit, seen, night):
    sources = [
        ("Яндекс.Карты", parse_yandex),
        ("2ГИС", parse_2gis),
    ]
    all_records = []
    tried_synonyms = False

    while len(all_records) < limit:
        new_this_cycle = 0

        for src_name, src_func in sources:
            if len(all_records) >= limit:
                break
            per_src = max(1, (limit - len(all_records)) // len(sources))
            print(f"\n {C.Cc}──  {src_name}: «{sphere} {city}»  ──{C.N}")
            print(f"     {C.D}Попытка собрать {per_src} зап. (нужно ещё {limit - len(all_records)}).{C.N}")

            try:
                records = src_func(page, city, sphere, per_src, seen, night)
                new_this_cycle += len(records)
                for rec in records:
                    all_records.append(rec)
                print(f"     {C.G}✓{C.N} +{len(records)} из {src_name}. Всего по комбинации: {len(all_records)}/{limit}")
            except Exception as e:
                print(f"     {C.R}⚠  {src_name}: {e}{C.N}")

            delay(night, 3, 6)

        if len(all_records) >= limit:
            break

        if new_this_cycle == 0 and tried_synonyms:
            print(f"     {C.Y}⚠  Данных больше нет.{C.N}")
            break

        if new_this_cycle == 0 and not tried_synonyms:
            q = sphere.lower()
            syns = []
            if any(w in q for w in ["маникюр","ногти","педикюр"]):
                syns = ["ногти", "ногтевая студия", "педикюр"]
            elif any(w in q for w in ["ресниц","бров"]):
                syns = ["брови", "лашмейкер", "ламинирование ресниц"]
            elif any(w in q for w in ["массаж","спа","spa"]):
                syns = ["спа", "антицеллюлитный массаж", "массаж"]
            elif any(w in q for w in ["салон","бьюти"]):
                syns = ["бьюти студия", "студия красоты", "косметолог"]
            elif any(w in q for w in ["косметолог"]):
                syns = ["косметология", "чистка лица"]
            else:
                syns = ["маникюр", "ногти", "салон красоты"]

            tried_synonyms = True

            for syn in syns:
                if len(all_records) >= limit:
                    break
                print(f"\n {C.D}⏳  Пауза, затем синоним: «{syn}»{C.N}")
                delay(night, 3, 5)

                for src_name, src_func in sources:
                    if len(all_records) >= limit:
                        break
                    need2 = limit - len(all_records)
                    per_src = max(1, need2 // len(sources))
                    print(f"\n {C.Cc}──  {src_name}: «{syn} {city}»  ──{C.N}")
                    print(f"     {C.D}Попытка собрать {per_src} зап.{C.N}")
                    try:
                        records = src_func(page, city, syn, per_src, seen, night)
                        new_this_cycle += len(records)
                        for rec in records:
                            all_records.append(rec)
                        print(f"     {C.G}✓{C.N} +{len(records)} из {src_name}. Всего: {len(all_records)}/{limit}")
                    except Exception as e:
                        print(f"     {C.R}⚠  {src_name}: {e}{C.N}")
                    delay(night, 3, 6)

            if new_this_cycle == 0:
                break

    return all_records


# ══════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════

def main():
    print()
    print(f" {C.BD}{C.Cc}╔{'═'*53}╗{C.N}")
    print(f" {C.BD}{C.Cc}║  Парсер Яндекс.Карты + 2ГИС{' '*(30)}║{C.N}")
    print(f" {C.BD}{C.Cc}╚{'═'*53}╝{C.N}")
    print()

    # ── Режим ──
    night_raw = ask(f" {C.B}🌙{C.N}  Ночной сбор (большие паузы)", "нет")
    night = night_raw.lower() in ("да", "yes", "д", "y")
    print(f"   {C.G}✓{C.N} Режим: {C.BD}{'НОЧНОЙ' if night else 'ОБЫЧНЫЙ'}{C.N}\n")

    # ── Города ──
    cities_raw = ask(f" {C.B}🏙{C.N}  Города (через запятую)")
    cities_input = [c.strip() for c in cities_raw.split(",") if c.strip()]
    cities = [resolve_city(c) for c in cities_input]
    for c in cities:
        print(f"   {C.G}✓{C.N} Город: {C.BD}{c}{C.N}")
    print()

    # ── Сферы ──
    spheres_raw = ask(f" {C.B}🔍{C.N}  Сфера/запрос (можно неск. через запятую)")
    spheres = [s.strip() for s in spheres_raw.split(",") if s.strip()]
    for s in spheres:
        print(f"   {C.G}✓{C.N} Поиск: {C.BD}{s}{C.N}")
    print()

    # ── Лимит ──
    if night:
        raw_limit = ask(f" {C.B}📊{C.N}  Сколько записей (макс. 500)", "500")
        try:
            total_limit = min(max(1, int(raw_limit)), 500)
        except:
            total_limit = 500
    else:
        raw_limit = ask(f" {C.B}📊{C.N}  Сколько записей", "50")
        try:
            total_limit = max(1, int(raw_limit))
        except:
            total_limit = 50
    print(f"   {C.G}✓{C.N} Лимит: {C.BD}{total_limit}{C.N}\n")

    # ── Файл ──
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_name = f"clients_{ts}.xlsx"
    fname = ask(f" {C.B}💾{C.N}  Имя Excel-файла", default_name)
    if not fname.endswith(".xlsx"):
        fname += ".xlsx"
    print(f"   {C.G}✓{C.N} Файл: {C.BD}{fname}{C.N}\n")

    # ── История ──
    history = read_history()
    combos = [(c, s) for c in cities for s in spheres]
    skip_combos = set()

    for city, sphere in combos:
        prev = find_history(city, sphere, history)
        if prev:
            yn = ask(
                f" {C.Y}⚠  Город «{city}», сфера «{sphere}» уже "
                f"искались {prev['date']} ({prev['count']} зап.)\n"
                f"     Пропустить?",
                "нет"
            )
            if yn.lower() in ("да", "yes", "д", "y"):
                skip_combos.add((city, sphere))

    # ── Запуск ──
    print(f"\n {C.G}{'─'*55}{C.N}")
    print(f" {C.BD}🚀  Запуск...{C.N}")
    print(f" {C.D}     Откроется окно Brave — оно нужно для 2ГИС.{C.N}")
    print(f" {C.D}     Просто сверните его (Alt+Tab), не закрывайте.{C.N}")
    if night:
        print(f" {C.D}     Ночной режим — паузы ×2.{C.N}")
    print(f" {C.G}{'─'*55}{C.N}\n")

    all_records = []
    seen_keys = set()
    stats = {}

    with sync_playwright() as pw:
        ctx = pw.chromium.launch_persistent_context(
            user_data_dir=BRAVE_PROFILE,
            executable_path=BRAVE_PATH,
            headless=False,
            no_viewport=True,
            locale="ru-RU",
            args=[
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-features=PasswordManagerLeakDetection",
                "--window-position=-3000,0",
            ],
            ignore_default_args=["--enable-automation"],
        )

        page = ctx.pages[0] if ctx.pages else ctx.new_page()
        page.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )

        for city, sphere in combos:
            if (city, sphere) in skip_combos:
                print(f" {C.Y}⏭  Пропущено: {city} / {sphere}{C.N}")
                continue

            remaining = total_limit - len(all_records)
            if remaining <= 0:
                break

            # Сколько комбинаций осталось (включая текущую)
            remaining_combos = sum(
                1 for c, s in combos
                if (c, s) not in skip_combos
                and combos.index((c, s)) >= combos.index((city, sphere))
            )
            per_combo = max(1, remaining // remaining_combos) if remaining_combos else remaining

            print(f"\n {C.BD}{C.Cc}{'═'*55}{C.N}")
            print(f" {C.BD}{C.Cc}📍  {city} / {sphere} (лимит: {per_combo}){C.N}")
            print(f" {C.BD}{C.Cc}{'═'*55}{C.N}")

            records = collect_city_sphere(page, city, sphere, per_combo, seen_keys, night)

            key = f"{city} / {sphere}"
            stats[key] = len(records)

            for rec in records:
                all_records.append(rec)

            append_history(city, sphere, len(records), fname)

            # Задержка между комбинациями
            if (city, sphere) != combos[-1]:
                print(f"\n {C.D}⏳  Пауза между комбинациями...{C.N}")
                if night:
                    time.sleep(random.uniform(30, 60))
                else:
                    time.sleep(random.uniform(5, 10))

        ctx.close()

    # ── Итог ──
    print(f"\n {C.G}{'═'*55}{C.N}")
    print(f" {C.BD}📊  Сбор завершён!{C.N}")
    print(f" {C.G}{'═'*55}{C.N}")

    if len(cities) > 1 or len(spheres) > 1:
        print(f"\n {C.BD}📈 Статистика:{C.N}")
        for combo, count in sorted(stats.items()):
            print(f"   {combo:40} {C.BD}{count}{C.N} зап.")
    print(f"\n   {C.BD}Всего:{C.N} {len(all_records)}/{total_limit} записей")

    if all_records:
        save_excel(all_records, fname)
        print(f"   {C.G}✅{C.N} Файл: {C.BD}{os.path.abspath(fname)}{C.N}")
        print(f"   {C.G}✅{C.N} История: {C.BD}{os.path.abspath(HISTORY_FILE)}{C.N}")

    print(f"\n {C.G}{'─'*55}{C.N}\n")


if __name__ == "__main__":
    main()
