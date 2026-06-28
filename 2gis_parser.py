"""
2ГИС Парсер — сбор данных о компаниях (маникюр, салоны красоты, ресницы, массаж)
===============================================================================

Как использовать:
  1. Установи зависимости:
       pip install openpyxl playwright
       python3 -m playwright install chromium

  2. Настрой конфиг внизу файла (города, поисковые запросы)

  3. Запусти:
       python3 2gis_parser.py

Скрипт собирает: название, телефон(ы), ссылку на 2ГИС, VK, Telegram
и сохраняет в Excel-файл.

ВАЖНО: При первом запуске Playwright скачает браузер Chromium (~150 МБ).
"""

import re
import os
import sys
import time
from datetime import date
from collections import OrderedDict

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("Ошибка: установите playwright — pip install playwright")
    print("Затем: python3 -m playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Ошибка: установите openpyxl — pip install openpyxl")
    sys.exit(1)


# ══════════════════════════════════════════════════════════
#  ПОЛЬЗОВАТЕЛЬСКИЙ КОНФИГ (меняй под свои задачи)
# ══════════════════════════════════════════════════════════

CONFIG = {
    # Города: 'английское_название_для_2ГИС': 'Русское название'
    "cities": OrderedDict([
        ("izhevsk",  "Ижевск"),
        ("votkinsk", "Воткинск"),
        ("glazov",   "Глазов"),
        ("sarapul",  "Сарапул"),
        ("mozhga",   "Можга"),
    ]),

    # Поисковые запросы — для каждого будет отдельный поиск
    "search_queries": [
        "маникюр",
        "ногти",
        "ресницы",
        "массаж",
    ],

    # Сколько раз прокрутить страницу вниз для подгрузки результатов
    # (2ГИС подгружает по ~12 карточек за раз)
    "scroll_times": 3,

    # Ждать после каждого скролла (сек), чтобы контент успел загрузиться
    "scroll_delay": 2.0,

    # Куда сохранить результат
    "output_file": "clients.xlsx",

    # Показывать браузер (False = без графического окна)
    "headless": True,
}


# ══════════════════════════════════════════════════════════
#  ИЗВЛЕЧЕНИЕ ДАННЫХ СО СТРАНИЦЫ ФИРМЫ
# ══════════════════════════════════════════════════════════

def parse_firm_page(page, firm_url, city_ru):
    """Открыть карточку фирмы, достать данные, вернуть словарь."""
    page.goto(firm_url, wait_until="domcontentloaded")
    time.sleep(1.5)

    # Ждём появления контента
    try:
        page.wait_for_selector("title", timeout=8000)
    except:
        pass

    html = page.content()

    # Название
    title_match = re.search(r"<title>(.*?) — 2ГИС</title>", html)
    name = title_match.group(1).strip() if title_match else "Неизвестно"

    # Телефоны (+7XXXXXXXXXX в HTML)
    raw_phones = set(re.findall(r"\+7(\d{10})", html))
    phones_list = []
    for p in sorted(raw_phones):
        phones_list.append(f"+7 ({p[:3]}) {p[3:6]}-{p[6:8]}-{p[8:10]}")
    phones_str = ", ".join(phones_list)

    # VK
    vk_match = re.findall(r"https?://vk\.com/[^\s\"\'<]+", html)
    vk = vk_match[0] if vk_match else ""

    # Telegram
    tg_match = re.findall(r"https?://t\.me/[^\s\"\'<&]+", html)
    tg = tg_match[0] if tg_match else ""

    # Определение категории по описанию
    desc_match = re.search(
        r'<meta[^>]*name="description"[^>]*content="([^"]*?)"', html
    )
    desc = (desc_match.group(1) or "").lower() if desc_match else ""
    cat = _detect_category(desc)

    comment = ""
    if vk:
        comment += f"VK: {vk} "
    if tg:
        comment += f"TG: {tg}"

    return {
        "category": cat,
        "city": city_ru,
        "name": name,
        "phones": phones_str,
        "link": firm_url,
        "comment": comment.strip(),
    }


def _detect_category(desc):
    """Определить категорию по тексту описания."""
    if any(w in desc for w in ["ногтев", "маникюр", "педикюр", "nail"]):
        return "Ногтевые услуги"
    if any(w in desc for w in ["ресниц"]):
        return "Ресницы/Брови"
    if any(w in desc for w in ["массаж", "spa"]):
        return "Массаж/СПА"
    if any(w in desc for w in ["косметолог"]):
        return "Косметология"
    if any(w in desc for w in ["парикмахер", "стрижк"]):
        return "Парикмахерская"
    return "Салон красоты"


def extract_firm_urls_from_page(page):
    """Достать все ссылки на карточки фирм из результатов поиска 2ГИС."""
    html = page.content()
    firm_ids = set(re.findall(r"/firm/(\d{13,20})", html))

    # Определяем город по текущему URL
    url = page.url
    city_match = re.search(r"2gis\.ru/(\w+)", url)
    city_en = city_match.group(1) if city_match else "izhevsk"

    return [f"https://2gis.ru/{city_en}/firm/{fid}" for fid in sorted(firm_ids)]


# ══════════════════════════════════════════════════════════
#  ПОИСК ПО ГОРОДУ
# ══════════════════════════════════════════════════════════

def search_city(browser, city_en, city_ru, queries, scroll_times, scroll_delay):
    """
    Выполнить поиск по городу по всем запросам.
    Возвращает список словарей с данными фирм.
    """
    all_firm_urls = set()
    ctx = browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/125.0.0.0 Safari/537.36"
        ),
        locale="ru-RU",
    )
    page = ctx.new_page()

    for query in queries:
        print(f"\n  🔍 Поиск: «{query}»")
        from urllib.parse import quote
        encoded = quote(query, safe="")
        url = f"https://2gis.ru/{city_en}/search/{encoded}"

        try:
            page.goto(url, wait_until="domcontentloaded", timeout=20000)
            time.sleep(2)
        except Exception as e:
            print(f"    Ошибка загрузки: {e}")
            continue

        # Скроллы для подгрузки результатов
        for s in range(scroll_times):
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(scroll_delay)

        urls = extract_firm_urls_from_page(page)
        print(f"    Найдено фирм: {len(urls)}")
        all_firm_urls.update(urls)

    ctx.close()

    # Теперь заходим в каждую фирму и собираем данные
    results = []
    seen_phones = set()

    print(f"\n  📄 Обработка карточек ({len(all_firm_urls)} шт)...")
    ctx2 = browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/125.0.0.0 Safari/537.36"
        ),
        locale="ru-RU",
    )
    page2 = ctx2.new_page()

    for i, firm_url in enumerate(sorted(all_firm_urls), 1):
        time.sleep(0.8)
        try:
            rec = parse_firm_page(page2, firm_url, city_ru)
            if not rec["phones"]:
                print(f"    [{i}/{len(all_firm_urls)}] ✗ {rec['name'][:30]:30} | нет телефона")
                results.append(rec)
                continue

            # Дедупликация по номерам
            phone_set = set(rec["phones"].split(", "))
            if phone_set and phone_set.issubset(seen_phones):
                print(f"    [{i}/{len(all_firm_urls)}] – {rec['name'][:30]:30} | дубль")
                continue
            seen_phones.update(phone_set)

            results.append(rec)
            print(f"    [{i}/{len(all_firm_urls)}] ✓ {rec['name'][:30]:30} | {rec['phones'][:25]}")
        except Exception as e:
            print(f"    [{i}/{len(all_firm_urls)}] ✗ ошибка: {firm_url[:50]} — {e}")

    ctx2.close()
    return results


# ══════════════════════════════════════════════════════════
#  ЗАПИСЬ В EXCEL
# ══════════════════════════════════════════════════════════

def write_excel(records, filename):
    """Сформировать Excel-файл с колонками как в clients.xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Клиенты"

    headers = [
        "Сфера", "Город", "Наименование", "Имя",
        "Телефон", "Правовой статус", "Комментарий",
        "Ответственный", "Ссылка", "В архиве", "Дата создания"
    ]

    h_font = Font(bold=True, size=11, color="FFFFFF")
    h_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = Alignment(horizontal="center")

    today_str = date.today().strftime("%d.%m.%Y")

    for i, rec in enumerate(records, 2):
        ws.cell(row=i, column=1, value=rec["category"])
        ws.cell(row=i, column=2, value=rec["city"])
        ws.cell(row=i, column=3, value=rec["name"])
        ws.cell(row=i, column=4, value="")
        ws.cell(row=i, column=5, value=rec["phones"])
        ws.cell(row=i, column=6, value="ИП")
        ws.cell(row=i, column=7, value=rec["comment"])
        ws.cell(row=i, column=8, value="")
        ws.cell(row=i, column=9, value=rec["link"])
        ws.cell(row=i, column=10, value="")
        ws.cell(row=i, column=11, value=today_str)

    widths = [18, 12, 48, 15, 38, 10, 70, 18, 85, 10, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(filename)
    print(f"\n✅ Файл сохранён: {os.path.abspath(filename)} ({len(records)} записей)")


# ══════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════

def main():
    cfg = CONFIG
    cities = cfg["cities"]
    queries = cfg["search_queries"]
    scroll_times = cfg["scroll_times"]
    scroll_delay = cfg["scroll_delay"]
    output = cfg["output_file"]
    headless = cfg["headless"]

    print("=" * 60)
    print("2ГИС Парсер (Playwright)")
    print("=" * 60)
    print(f"Города: {', '.join(cities.values())}")
    print(f"Запросы: {', '.join(queries)}")
    print(f"Скроллов: {scroll_times}")
    print(f"Headless: {headless}")
    print("=" * 60)

    all_records = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=headless)

        for city_en, city_ru in cities.items():
            print(f"\n📍 {city_ru} ({city_en})")
            records = search_city(
                browser, city_en, city_ru, queries, scroll_times, scroll_delay
            )
            print(f"  Итого по городу: {len(records)} записей")
            all_records.extend(records)

        browser.close()

    print(f"\n{'=' * 60}")
    print(f"📊 Всего собрано: {len(all_records)} записей")

    if all_records:
        write_excel(all_records, output)
    else:
        print("Нет данных для сохранения.")


if __name__ == "__main__":
    main()
